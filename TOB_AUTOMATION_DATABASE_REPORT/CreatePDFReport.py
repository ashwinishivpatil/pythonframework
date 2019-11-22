from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle, TA_CENTER,TA_LEFT
from reportlab.lib.units import inch, mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, SimpleDocTemplate, Spacer,TableStyle
from reportlab.lib import colors
import openpyxl
from reportlab.pdfbase import pdfdoc
import os
#import urllib2_file
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Image
testpassCount = 0
testfailCount = 0
import os
########################################################################

fileName = ""
chartName = ""

class Test(object):
    """"""

    # ----------------------------------------------------------------------
    def __init__(self):
        """Constructor"""
        self.width, self.height = letter
        self.styles = getSampleStyleSheet()
        self.leading = 24
    # ----------------------------------------------------------------------
    def coord(self, x, y, unit=1):
        """
        http://stackoverflow.com/questions/4726011/wrap-text-in-a-table-reportlab
        Helper class to help position flowables in Canvas objects
        """
        x, y = x * unit, self.height - y * unit
        return x, y

    # ----------------------------------------------------------------------
    def run(self):
        """
        Run the report
        """
        self.doc = SimpleDocTemplate(self.pdfname)
        self.story = [Spacer(1, 2.5 * inch)]
        self.createLineItems()

        self.doc.build(self.story, onFirstPage=self.createDocument)


    # ----------------------------------------------------------------------
    def createDocument(self, canvas, doc):
        """
        Create the document
        """
        self.c = canvas
        normal = self.styles["Normal"]

        header_text = "<h1><b><i></i></b></h1>"
        centered =  ParagraphStyle('parrafos',alignment = TA_LEFT, fontSize = 15,
                           fontName="HELVETICA")
        p = Paragraph(header_text, centered)
        p.wrapOn(self.c, self.width, self.height)
        p.drawOn(self.c, *self.coord(100, 12, mm))

        ptext = """&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                   <U>Automation Test Scripts For TOB Payer Back Bone                </U>         """

        p = Paragraph(ptext, style=centered)

        p.wrapOn(self.c, self.width - 50, self.height)
        p.drawOn(self.c, 30, 700)



    # ----------------------------------------------------------------------
    def createLineItems(self):
        """
        Create the line items
        """
        path = os.getcwd()
        os.chdir(path)

        im = Image(self.chartName+".png", 7 * inch, 7 * inch)
        self.story.append(im)

        """ totalData = []
        line_data2 = ["Total Number of Pass"+str(testpassCount),"Total Number of Fail"+str(testfailCount),
                      "", ""]
        centered1 = ParagraphStyle(name="left", alignment=TA_LEFT)
        format_line_data = []
        for item in line_data2:
            ptext = "<font size=%s><B>%s</B></font>" % (12, item)
            p = Paragraph(ptext, centered1)

            format_line_data.append(p)

        totalData.append(format_line_data)
        table = Table(totalData, colWidths=[300, 100, 100, 100])
        self.story.append(table)"""


        text_data = ["TESTCASE STATUS", "TESTCASE NAME ", "TEST FAIL DESCRIPTION",
                     "EXPECTED RESULT"]
        d = []
        font_size = 10
        centered =  ParagraphStyle('parrafos',alignment = TA_LEFT, fontSize = 8,
                           fontName="Helvetica",textColor = 'black')

        for text in text_data:
            ptext = "<b><font size=%s color='darkblue'>%s</font></b>" % (font_size, text)
            p1 = Paragraph(ptext, centered)

            d.append(p1)

        dataTitle = [d]

        #line_num = 1

        table = Table(dataTitle, colWidths=[50, 150, 150, 150])
        table.setStyle(TableStyle(
            [('LINEABOVE', (0, 0), (-1, 0), 2, colors.darkblue),
             ('LINEABOVE', (0, 1), (-1, -1), 0.25, colors.green),
             ('LINEBELOW', (0, -1), (-1, -1), 2, colors.darkblue),
             ('ALIGN', (1, 1), (-1, -1), 'RIGHT'), ('VALIGN', (0, -1), (-1, -1), 'BOTTOM'),
             ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
             ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]
        ))

        wb = openpyxl.load_workbook(self.fileName)
        sheetsNames = wb.sheetnames
        sheetsNames.remove("Sheet")
        print("sheetsNames",sheetsNames)
        sheetsNames = ["Entity",
                        "EntityFormulary",
                        "DrugFormularyStatus",
                        "DrugFormularyRestrictions",
                        "AccountRollup",
                        "ChannelRollup",
                        "Drug",
                        "PlanProduct",
                        "Person",
                        "PBMServices",
                        "KeyContact",
                        "IMSBridge",
                        "Formulary",
                        "EntityPerson",
                        "EntityProduct",
                        "EntityProductFormulary",
                        "EntitySubChannel",
                       "PredominantFormularyData"
                        ]
        for j in range(sheetsNames.__len__()):
            formatted_line_data = []
            data = []
            line_num = 1

            sheet = wb[str(sheetsNames[j])]

            data1 = []
            line_data1 = [" "+str(sheetsNames[j].capitalize())+"  Validation", "",
             "", ""]
            centered1 = ParagraphStyle(name="left", alignment=TA_CENTER)
            for item in line_data1:
                ptext = "<font size=%s><B>%s</B></font>" % (12, item)
                p = Paragraph(ptext, centered1)

                formatted_line_data.append(p)

            data1.append(formatted_line_data)
            table = Table(data1, colWidths=[450, 50, 50, 50])
            self.story.append(table)
            formatted_line_data = []
            for i in range(1, sheet.max_row + 1):
                #line_data = [str(line_num), "Check the columns for IMSBridge tab", "The Column is need present{'IMSPlanName', 'RxType', 'EntityID', 'IMSId', 'DisplayName\n', 'SubChannel'}"
                 #            ,"PASS"] sheet.cell(row=i,column=1).value
                flag = 0
                temp = 0
                if(str(sheet.cell(row=i,column=1).value).__contains__("F")):
                    flag = 1
                line_data = [str(sheet.cell(row=i,column=1).value),str(sheet.cell(row=i,column=2).value),str(sheet.cell(row=i,column=3).value),str(sheet.cell(row=i,column=4).value)]

                for item in line_data:
                    if(flag ==1):
                        ptext1 = "<font size=%s color='red'>%s</font>" % (font_size - 1, item)

                    else:
                        if(i==1):
                         ptext1 = "<font size=%s><B>%s</B></font>" % (font_size + 1, item)
                        else:
                         ptext1 = "<font size=%s>%s</font>" % (font_size - 1, item)
                    p = Paragraph(ptext1, centered)

                    formatted_line_data.append(p)

                data.append(formatted_line_data)
                formatted_line_data = []
                line_num += 1

            table = Table(data, colWidths=[50, 150,150,150])
            table.setStyle(TableStyle(
        [('LINEABOVE', (0,0), (-1,0), 2, colors.darkblue),
        ('LINEABOVE', (0,1), (-1,-1), 0.25, colors.green),
        ('LINEBELOW', (0,-1), (-1,-1), 2, colors.darkblue),
        ('ALIGN', (1,1), (-1,-1), 'RIGHT'),('VALIGN',(0,-1),(-1,-1),'BOTTOM'),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),]
    ))


            self.story.append(table)
            self.story.append(Spacer(1, 0.25 * inch))





def CreatePDF(excelfilePath,chartFileName,pdfFileName):

        fileName = excelfilePath
        chartName = chartFileName
        t = Test()
        t.fileName = excelfilePath
        t.chartName = chartFileName
        t.pdfname = pdfFileName
        t.run()
